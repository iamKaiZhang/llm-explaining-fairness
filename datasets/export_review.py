"""Export datasets/llm_eval_dataset/ to one Excel workbook for human review.

    .venv/bin/python datasets/export_review.py

Writes datasets/llm_eval_dataset_review.xlsx: an Overview sheet (one row per
question set) plus one sheet per set with a row per question — main form,
paraphrases side by side, gold answer(s), and provenance details. Re-run
after regenerating the dataset.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATASET_DIR = Path(__file__).resolve().parent / "llm_eval_dataset"
OUT_PATH = Path(__file__).resolve().parent / "llm_eval_dataset_review.xlsx"
# optional: {question id -> proposed rewrite}; adds a review column when present
PROPOSALS_PATH = Path(__file__).resolve().parent / "proposed_questions.json"

SET_FILES = [
    "coverage_disparity", "coverage", "benefit_inequality",
    "group_coverage_gap", "demographic_parity", "worst_group",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


def _fmt(value) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return "" if value is None else str(value)


def _write_sheet(ws, headers: list[str], rows: list[list], widths: list[int]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for row in rows:
        ws.append([_fmt(v) if not isinstance(v, (int, float, bool)) else v
                   for v in row])
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    ws.freeze_panes = "A2"


def main() -> None:
    sets = [json.loads((DATASET_DIR / f"{n}.json").read_text()) for n in SET_FILES]
    proposals: dict[str, str] = {}
    if PROPOSALS_PATH.exists():
        proposals = json.loads(PROPOSALS_PATH.read_text())
        print(f"including {len(proposals)} proposed rewrites from {PROPOSALS_PATH.name}")
    wb = Workbook()

    # --- overview sheet -------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    rows = []
    for s in sets:
        n_para = sum(len(q["paraphrases"]) for q in s["questions"])
        rows.append([
            s["set_id"], s["expected_tool"], len(s["questions"]), n_para,
            len(s["questions"]) + n_para, s["template"], s["definition"],
        ])
    _write_sheet(
        ws,
        ["Set", "Expected tool", "Questions", "Paraphrases", "Test inputs",
         "Template", "Definition"],
        rows,
        [24, 22, 10, 12, 11, 45, 80],
    )

    # --- one sheet per set ----------------------------------------------
    max_para = max(len(q["paraphrases"]) for s in sets for q in s["questions"])
    for s in sets:
        ws = wb.create_sheet(s["set_id"].removesuffix("-v1")[:31])
        has_metric = any("expected_metric" in q for q in s["questions"])
        has_choice = any("expected_choice" in q for q in s["questions"])
        has_bool = any("answer_bool" in q for q in s["questions"])

        headers = ["ID", "Params", "Question (main form)"]
        if proposals:
            headers.append("Proposed question (Codex)")
        headers += [f"Paraphrase {i + 1}" for i in range(max_para)]
        headers += ["Answer", "Tolerance"]
        if has_metric:
            headers.append("Expected metric")
        if has_choice:
            headers.append("Expected worst group")
        if has_bool:
            headers.append("Verdict (bool)")
        headers.append("Answer details")

        rows = []
        for q in s["questions"]:
            paras = q["paraphrases"] + [""] * (max_para - len(q["paraphrases"]))
            row = [q["id"], json.dumps(q["params"]), q["question"]]
            if proposals:
                row.append(proposals.get(q["id"], ""))
            row += [*paras, q["answer"], q["tolerance"]]
            if has_metric:
                row.append(q.get("expected_metric", ""))
            if has_choice:
                row.append(q.get("expected_choice", ""))
            if has_bool:
                row.append(q.get("answer_bool", ""))
            row.append(json.dumps(q["answer_details"]))
            rows.append(row)

        widths = [8, 30, 45] + ([45] if proposals else []) + [45] * max_para + [10, 10]
        widths += [15] * (has_metric + has_choice + has_bool) + [50]
        _write_sheet(ws, headers, rows, widths)

    wb.save(OUT_PATH)
    n_q = sum(len(s["questions"]) for s in sets)
    n_p = sum(len(q["paraphrases"]) for s in sets for q in s["questions"])
    print(f"wrote {OUT_PATH}  ({len(sets)} sets, {n_q} questions, {n_p} paraphrases)")


if __name__ == "__main__":
    main()
